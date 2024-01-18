# To extract hyperlinks, we need to use a different approach since pandas doesn't directly support hyperlink extraction.
# We will use openpyxl to extract the hyperlinks

import openpyxl
import pandas as pd

# Load the workbook and the active sheet
workbook = openpyxl.load_workbook('/Users/jorge/Documents/USA_Stations_Q_v0.xlsx')
sheet = workbook.active

# Prepare a list to store the hyperlinks
hyperlinks = []

# Assuming the hyperlinks are in the second column (Site Number)
for row in range(2, sheet.max_row + 1):  # Starting from 2 to skip the header
    cell = sheet.cell(row, 2)  # Second column for Site Number
    if cell.hyperlink:
        hyperlinks.append(cell.hyperlink.target)
    else:
        hyperlinks.append(None)

# Now, add the hyperlinks to the dataframe
df = pd.DataFrame()
df['Hyperlink'] = hyperlinks

print(df)

# Check the first few rows of the updated dataframe
df.head()
